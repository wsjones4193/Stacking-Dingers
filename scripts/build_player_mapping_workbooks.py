"""
Build player mapping workbooks — one Excel file per tournament year.

Each workbook has three tabs:
  1. underdog_players  — unique players from the Underdog CSV
                         (player_name, underdog_player_id, position, ud_team)
  2. mlb_players       — all MLB players for that season from Stats API
                         (player_name, mlb_id, position, mlb_team)
  3. mapping           — fuzzy-matched join of the two
                         (ud_player_name, underdog_player_id, ud_position,
                          mlb_player_name, mlb_id, mlb_team, match_score,
                          match_type)

Output: data/mapping/{season}_player_mapping.xlsx

Usage:
  python scripts/build_player_mapping_workbooks.py --seasons 2022 2023 2024 2025
"""

from __future__ import annotations

import argparse
import logging
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd
import statsapi
from rapidfuzz import fuzz, process

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

CSV_DIR = Path("data/csv")
OUT_DIR = Path("data/mapping")

# Fuzzy match score thresholds
EXACT_THRESHOLD = 100
HIGH_THRESHOLD = 90    # confident auto-match
LOW_THRESHOLD = 70     # possible match — flag for review


# ---------------------------------------------------------------------------
# Step 1 — Underdog players from CSV
# ---------------------------------------------------------------------------

def get_underdog_players(season: int) -> pd.DataFrame:
    """
    Extract unique players from the Underdog season CSV.
    Returns: player_name, underdog_player_id, position
    """
    path = CSV_DIR / f"{season}.csv"
    if not path.exists():
        logger.warning(f"No CSV found for {season} at {path}")
        return pd.DataFrame()

    df = pd.read_csv(path, usecols=["player_name", "underdog_player_id", "position"])

    # 2022 data has no underdog_player_id — use player_name as key
    if df["underdog_player_id"].isna().all():
        unique = (
            df.groupby("player_name")
            .agg(position=("position", "first"))
            .reset_index()
        )
        unique["underdog_player_id"] = ""
    else:
        unique = (
            df.dropna(subset=["underdog_player_id"])
            .groupby("underdog_player_id")
            .agg(
                player_name=("player_name", "first"),
                position=("position", "first"),
            )
            .reset_index()
        )
        unique.columns = ["underdog_player_id", "player_name", "position"]

    unique = unique.sort_values("player_name").reset_index(drop=True)
    unique = unique[["player_name", "underdog_player_id", "position"]]
    logger.info(f"  Underdog {season}: {len(unique)} unique players")
    return unique


# ---------------------------------------------------------------------------
# Step 2 — MLB players from Stats API
# ---------------------------------------------------------------------------

def get_mlb_players(season: int) -> pd.DataFrame:
    """
    Fetch all MLB players for a season with team info.
    Returns: player_name, mlb_id, position, mlb_team
    """
    logger.info(f"  Fetching MLB roster for {season}...")
    try:
        data = statsapi.get("sports_players", {"season": season, "gameType": "R"})
        people = data.get("people", [])
    except Exception as e:
        logger.error(f"  Failed to fetch MLB players for {season}: {e}")
        return pd.DataFrame()

    rows = []
    for p in people:
        mlb_id = p.get("id")
        name = p.get("fullName", "")
        pos = p.get("primaryPosition", {}).get("abbreviation", "")
        # currentTeam may not be present for all players
        team = p.get("currentTeam", {}).get("name", "")
        if mlb_id and name:
            rows.append({
                "player_name": name,
                "mlb_id": mlb_id,
                "position": pos,
                "mlb_team": team,
            })

    df = pd.DataFrame(rows).sort_values("player_name").reset_index(drop=True)
    logger.info(f"  MLB {season}: {len(df)} players")
    return df


# ---------------------------------------------------------------------------
# Step 3 — Fuzzy match
# ---------------------------------------------------------------------------

def normalize_name(name: str) -> str:
    """Normalize a player name for matching: lowercase, strip punctuation."""
    import re
    name = name.lower().strip()
    # Remove suffixes: Jr., Sr., II, III, IV
    name = re.sub(r"\b(jr\.?|sr\.?|ii|iii|iv)\b", "", name)
    # Remove punctuation except spaces
    name = re.sub(r"[^a-z\s]", "", name)
    # Collapse whitespace
    name = re.sub(r"\s+", " ", name).strip()
    return name


def fuzzy_match(ud_players: pd.DataFrame, mlb_players: pd.DataFrame) -> pd.DataFrame:
    """
    For each Underdog player, find the best MLB match using fuzzy name matching.
    Returns a DataFrame with both sides joined + match metadata.
    """
    # Build normalized name lookup for MLB players
    mlb_names = mlb_players["player_name"].tolist()
    mlb_names_norm = [normalize_name(n) for n in mlb_names]

    rows = []
    for _, ud_row in ud_players.iterrows():
        ud_name = ud_row["player_name"]
        ud_name_norm = normalize_name(ud_name)

        # Try exact normalized match first
        exact_idx = None
        for i, n in enumerate(mlb_names_norm):
            if n == ud_name_norm:
                exact_idx = i
                break

        if exact_idx is not None:
            mlb_row = mlb_players.iloc[exact_idx]
            rows.append({
                "ud_player_name": ud_name,
                "underdog_player_id": ud_row["underdog_player_id"],
                "ud_position": ud_row["position"],
                "mlb_player_name": mlb_row["player_name"],
                "mlb_id": mlb_row["mlb_id"],
                "mlb_team": mlb_row["mlb_team"],
                "mlb_position": mlb_row["position"],
                "match_score": 100,
                "match_type": "exact",
                "needs_review": "",
            })
            continue

        # Fuzzy match against normalized names
        result = process.extractOne(
            ud_name_norm,
            mlb_names_norm,
            scorer=fuzz.token_sort_ratio,
        )

        if result is None:
            rows.append({
                "ud_player_name": ud_name,
                "underdog_player_id": ud_row["underdog_player_id"],
                "ud_position": ud_row["position"],
                "mlb_player_name": "",
                "mlb_id": None,
                "mlb_team": "",
                "mlb_position": "",
                "match_score": 0,
                "match_type": "no_match",
                "needs_review": "YES",
            })
            continue

        best_name_norm, score, idx = result
        mlb_row = mlb_players.iloc[idx]

        if score >= HIGH_THRESHOLD:
            match_type = "fuzzy_high"
            needs_review = ""
        elif score >= LOW_THRESHOLD:
            match_type = "fuzzy_low"
            needs_review = "REVIEW"
        else:
            match_type = "no_match"
            needs_review = "YES"

        rows.append({
            "ud_player_name": ud_name,
            "underdog_player_id": ud_row["underdog_player_id"],
            "ud_position": ud_row["position"],
            "mlb_player_name": mlb_row["player_name"],
            "mlb_id": mlb_row["mlb_id"] if score >= LOW_THRESHOLD else None,
            "mlb_team": mlb_row["mlb_team"] if score >= LOW_THRESHOLD else "",
            "mlb_position": mlb_row["position"] if score >= LOW_THRESHOLD else "",
            "match_score": score,
            "match_type": match_type,
            "needs_review": needs_review,
        })

    result_df = pd.DataFrame(rows)

    # Sort: unmatched first, then low-confidence, then high-confidence
    sort_order = {"no_match": 0, "fuzzy_low": 1, "fuzzy_high": 2, "exact": 3}
    result_df["_sort"] = result_df["match_type"].map(sort_order)
    result_df = result_df.sort_values(["_sort", "ud_player_name"]).drop(columns="_sort").reset_index(drop=True)

    n_exact = (result_df["match_type"] == "exact").sum()
    n_high = (result_df["match_type"] == "fuzzy_high").sum()
    n_low = (result_df["match_type"] == "fuzzy_low").sum()
    n_none = (result_df["match_type"] == "no_match").sum()
    logger.info(f"  Match results: {n_exact} exact, {n_high} fuzzy-high, {n_low} fuzzy-low (review), {n_none} unmatched")

    return result_df


# ---------------------------------------------------------------------------
# Step 4 — Write Excel workbook
# ---------------------------------------------------------------------------

def write_workbook(
    season: int,
    ud_players: pd.DataFrame,
    mlb_players: pd.DataFrame,
    mapping: pd.DataFrame,
) -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / f"{season}_player_mapping.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        ud_players.to_excel(writer, sheet_name="underdog_players", index=False)
        mlb_players.to_excel(writer, sheet_name="mlb_players", index=False)
        mapping.to_excel(writer, sheet_name="mapping", index=False)

        # Auto-size columns and highlight rows needing review
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter

        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]

            # Auto-size columns
            for col in ws.columns:
                max_len = max(
                    (len(str(cell.value)) if cell.value is not None else 0)
                    for cell in col
                )
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

            # Highlight mapping sheet rows
            if sheet_name == "mapping":
                review_col = None
                for cell in ws[1]:
                    if cell.value == "needs_review":
                        review_col = cell.column
                        break

                if review_col:
                    for row in ws.iter_rows(min_row=2):
                        review_val = row[review_col - 1].value
                        if review_val == "YES":
                            for cell in row:
                                cell.fill = red
                        elif review_val == "REVIEW":
                            for cell in row:
                                cell.fill = yellow

    logger.info(f"  Written: {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_season(season: int) -> None:
    logger.info(f"=== Season {season} ===")
    ud_players = get_underdog_players(season)
    if ud_players.empty:
        return

    mlb_players = get_mlb_players(season)
    if mlb_players.empty:
        return

    mapping = fuzzy_match(ud_players, mlb_players)
    write_workbook(season, ud_players, mlb_players, mapping)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--seasons", nargs="+", type=int, default=[2022, 2023, 2024, 2025])
    args = parser.parse_args()

    for season in args.seasons:
        build_season(season)
        time.sleep(1)  # brief pause between seasons

    logger.info("Done. Review files in data/mapping/")


if __name__ == "__main__":
    main()
